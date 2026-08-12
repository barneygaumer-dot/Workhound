from io import BytesIO
from openpyxl import load_workbook
from .base import ImportedDocument, WorkCandidate
from .csv_adapter import _map

def parse_xlsx(data, filename):
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    doc = ImportedDocument(source_type="xlsx", title=filename)
    if not rows: return doc
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    mapping = _map(headers)
    doc.metadata["mapping"] = mapping
    title_col = mapping.get("title")
    if not title_col: return doc
    idx = {h:i for i,h in enumerate(headers)}
    def val(row, field, default=""):
        h = mapping.get(field)
        if not h or h not in idx or idx[h] >= len(row): return default
        v = row[idx[h]]
        return default if v is None else str(v).strip()
    for n,row in enumerate(rows[1:],2):
        title = val(row,"title")
        if not title: continue
        try: p=int(float(val(row,"progress_percent","0").replace("%","")))
        except: p=0
        doc.candidates.append(WorkCandidate(
            title=title, description=val(row,"description",val(row,"notes")),
            owner=val(row,"owner"), priority=val(row,"priority","Medium"),
            status=val(row,"status","New"), category=val(row,"category"),
            progress_percent=max(0,min(100,p)), source_section=f"row {n}",
            source_text=str(row), key=f"xlsx-{n}"
        ))
    return doc
